#!/usr/bin/python
import csv
import numpy as np
import matplotlib.pyplot as plt
from numpy.linalg import inv
from openpyxl import load_workbook

# parameters
dx = 0.05    # space resolution [m]
dt = 1.0     # time resolution [day]
D  = 2.5e-5  # dispersion coefficient [m2/day]
u  = 0.005   # advection speed [m/day]
N  = 25     # number of dx

cfeed = 1.


# Global settings
tprint = [50, 100]
xmin, xmax = 0.0, dx*N
x_pos = np.linspace(xmin, xmax, N)  # discretized x
# plot settings
linstyles = ['-', '--', 'o-']
colors = ['b', 'k', 'r']

def main():
    with open('compare.csv', 'w') as csvfile:
        fieldnames = ['method', 'dt', 'dx', 't', 'change in storage',
                      'flux out of the domain', 'diff', 'total_error']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()

        # Question 1 (b)
        c_explicit_upstream = finite_difference_c(dx, dt, u, D, tprint, writer,
                                                  time_method="explicit",
                                                  space_method="upstream")
        c_explicit_central = finite_difference_c(dx, dt, u, D, tprint, writer,
                                                 time_method="explicit",
                                                 space_method="central")
        plot_c(tprint, c_explicit_central, c_explicit_upstream, "explicit")
        saveresult(tprint, c_explicit_central, c_explicit_upstream, "explicit")

        # Question 1 (c)
        # dt_unstable = 1./(2.*D/(dx**2)+u/dx) + 0.1
        dt_unstable = 2
        tprint_2 = [10*dt_unstable, 20*dt_unstable, 30*dt_unstable, 40*dt_unstable]
        c_explicit_upstream_unstable = finite_difference_c(dx, dt_unstable, u, D, tprint_2, writer,
                                                           time_method="explicit",
                                                           space_method="upstream")
        # c_explicit_upstream_unstable2 = finite_difference_c(dx, dt_unstable, u, D, tprint_2, writer,
        #                                                     time_method="explicit",
        #                                                     space_method="upstream_1st")

        plot_c_unstable(tprint_2, c_explicit_upstream_unstable)

        # Question 2 (b)
        c_CN_upstream = finite_difference_c(dx, dt, u, D, tprint, writer,
                                            time_method="Crank-Nicolson",
                                            space_method="upstream")
        c_CN_central = finite_difference_c(dx, dt, u, D, tprint, writer,
                                           time_method="Crank-Nicolson",
                                           space_method="central")
        plot_c(tprint, c_CN_central, c_CN_upstream, "Crank-Nicolson")
        saveresult(tprint, c_CN_central, c_CN_upstream, "Crank-Nicolson")

        plt.show()

def tridiag(a, b, c, n, k1=-1, k2=0, k3=1):
    a, b, c = a*np.ones(n-abs(k1)), b*np.ones(n-abs(k2)), c*np.ones(n-abs(k3))
    return np.matrix(np.diag(a, k1) + np.diag(b, k2) + np.diag(c, k3))

def ifin(t, tprint):
    for tcheck in tprint:
        if np.isclose(t, tcheck):
            return True
    return False

def analytical_c(n, x, t):
    h_n = lambda n, x, t: 1.0/n**2 * np.exp(-(n**2)*(np.pi**2)*t) * np.sin(n*np.pi/2) * np.sin(n*np.pi*x)
    h_n_num = 0
    for i in xrange(n):
        i_n = 2*(i+1) - 1
        h_n_num += h_n(i_n, x, t)
    return h_n_num * 8 / (np.pi**2)

def prepare_matrices(dx, dt, u, D, N, method_name):
    SD = D*dt/(dx**2.)
    SU = u*dt/dx
    alpha = dx*u/D
    #alpha=1.1
    if method_name == "Explicit (upstream)":
        A = np.identity(N)
        B = tridiag(SD+SU, -(2*SD+SU-1), SD, N)
        B[0, 0] = 1-(alpha+1)*SD-alpha*SU
        b = np.zeros(N)
        b[0] = alpha*(SD+SU)*cfeed
        A, B, b = np.matrix(A), np.matrix(B), np.matrix(b).T
    elif method_name == "Explicit (upstream_1st)":
        A = np.identity(N-1)
        B = tridiag(SD+SU, -(2*SD+SU-1), SD, N-1)
        b = np.zeros(N-1)
        b[0] = (SD+SU)*cfeed
        A, B, b = np.matrix(A), np.matrix(B), np.matrix(b).T
    elif method_name == "Explicit (central)":
        A = np.identity(N)
        B = tridiag(SD+SU/2., 1-2*SD, SD-SU/2., N)
        B[0, 0] = 1-2*(alpha+1)*SD-alpha*SU
        B[0, 1] = 2.*SD
        b = np.zeros(N)
        b[0] = alpha*(2.*SD+SU)*cfeed
        A, B, b = np.matrix(A), np.matrix(B), np.matrix(b).T
    elif method_name == "Crank-Nicolson (upstream)":
        Ad = tridiag(-SD/2., 1.+SD, -SD/2., N)
        Ad[0, 0] = 1.+(alpha+1.)/2.*SD
        Aa = tridiag(-SU/2., SU/2., 0., N)
        Aa[0, 0] = alpha*SU/2.
        A = Aa + Ad
        Bd = tridiag(SD/2., 1.-SD, SD/2., N)
        Bd[0, 0] = 1.-(alpha+1.)/2.*SD
        Ba = tridiag(SU/2., -SU/2., 0., N)
        Ba[0, 0] = -alpha*SU/2.
        B = Bd + Ba
        b = np.zeros(N)
        b[0] = alpha*(SD+SU)*cfeed
        A, B, b = np.matrix(A), np.matrix(B), np.matrix(b).T
    elif method_name == "Crank-Nicolson (central)":
        Ad = tridiag(-SD/2., 1.+SD, -SD/2., N)
        Ad[0, 0] = 1.+(alpha+1.)*SD
        Ad[0, 1] = -SD
        Aa = tridiag(-SU/4., 0, SU/4., N)
        Aa[0, 0] = alpha*SU/2.
        Aa[0, 1] = 0.
        A = Aa + Ad
        Bd = tridiag(SD/2., 1.-SD, SD/2., N)
        Bd[0, 0] = 1.-(alpha+1)*SD
        Bd[0, 1] = SD
        Ba = tridiag(SU/4., 0, -SU/4., N)
        Ba[0, 0] = -alpha*SU/2.
        Ba[0, 1] = 0.
        B = Bd + Ba
        b = np.zeros(N)
        b[0] = alpha*(2.*SD+SU)*cfeed
        A, B, b = np.matrix(A), np.matrix(B), np.matrix(b).T
    elif method_name == "Crank-Nicolson (compare)":
        Ad = tridiag(-SD/2., 1.+SD, -SD/2., N)
        Aa = tridiag(-SU/4., 0, SU/4., N)
        A = Aa + Ad
        Bd = tridiag(SD/2., 1.-SD, SD/2., N)
        Ba = tridiag(SU/4., 0, -SU/4., N)
        B = Bd + Ba
        b = np.zeros(N)
        b[0] = (SD+SU)*cfeed
        A, B, b = np.matrix(A), np.matrix(B), np.matrix(b).T
    else:
        raise Exception("Unknown method")

    return [A, B, b]

def finite_difference_c(dx, dt, u, D, tprint, writer, tmax=101.0,
                        space_method="upstream", time_method="explicit"):

    # discretized x
    x_pos = np.linspace(xmin, xmax, N+1)
    # initial conditions
    c0 = np.zeros(N)

    # Script D & Script U
    SD = D*dt/(dx**2.)
    SU = u*dt/dx

    # x value
    c = [c0]
    c_print = []

    # build up the matrices
    if (space_method == "upstream") and (time_method == "explicit"):
        method_name = "Explicit (upstream)"
        [A, B, b] = prepare_matrices(dx, dt, u, D, N, method_name)
        Ainv = inv(A)
    elif (space_method == "upstream_1st") and (time_method == "explicit"):
        method_name = "Explicit (upstream_1st)"
        [A, B, b] = prepare_matrices(dx, dt, u, D, N, method_name)
        Ainv = inv(A)
    elif (space_method == "central") and (time_method == "explicit"):
        method_name = "Explicit (central)"
        [A, B, b] = prepare_matrices(dx, dt, u, D, N, method_name)
        Ainv = inv(A)
    elif (space_method == "upstream") and (time_method == "Crank-Nicolson"):
        method_name = "Crank-Nicolson (upstream)"
        [A, B, b] = prepare_matrices(dx, dt, u, D, N, method_name)
        Ainv = inv(A)
    elif (space_method == "central") and (time_method == "Crank-Nicolson"):
        method_name = "Crank-Nicolson (central)"
        [A, B, b] = prepare_matrices(dx, dt, u, D, N, method_name)
        Ainv = inv(A)
    else:
        raise Exception("Unknown method")

    print method_name
    print A
    print B
    print b
    # loop
    t = 0.0
    c_old = np.matrix(c0).T
    while (t <= tmax):
        # calculate x values at time step k+1
        c_new = Ainv * (B * c_old + b)
        # append x_new to x
        c.append(c_new)
        # save the values to be printed in c_print
        if ifin(t, tprint):
            c_print.append(c_new)
        # update t
        t += dt
        # assign x_new to x_old
        c_old = c_new

    return c_print

def plot_c(tprint, c_central, c_upstream, method):

    pos_anal, val_anal = [], []

    # load the analytical solution
    # T = 50 days
    wbsheet = load_workbook('cp2_t_50_3rd.xlsx')['Sheet1'].columns
    pos_anal.append([cell.value for cell in wbsheet[1][1:]])
    val_anal.append([cell.value for cell in wbsheet[2][1:]])

    # T = 100 days
    wbsheet = load_workbook('cp2_t_100_3rd.xlsx')['Sheet1'].columns
    pos_anal.append([cell.value for cell in wbsheet[1][1:]])
    val_anal.append([cell.value for cell in wbsheet[2][1:]])

    for i in xrange(len(tprint)):
        fig, ax = plt.subplots()
        ax.plot(x_pos, c_upstream[i], '--', label="upstream")
        ax.plot(x_pos, c_central[i], 'o-', label="central")
        ax.plot(pos_anal[i], val_anal[i], label="analytical")
        ax.legend(loc="upper right")
        ax.set_ylim([0, 1.5])
        ax.set_ylabel('c(x)')
        ax.set_xlabel('x')
        ax.set_title('Method: %s; Time: %s' % (method, str(tprint[i])))
        plt.savefig(method + "_" + str(tprint[i]) + ".png")

def plot_c_unstable(tprint, c_unstable):
    fig, axaar = plt.subplots(2, 2)
    c_unstable[0]
    axaar[0, 0].plot(x_pos, c_unstable[0])
    axaar[0, 1].plot(x_pos, c_unstable[1])
    axaar[1, 0].plot(x_pos, c_unstable[2])
    axaar[1, 1].plot(x_pos, c_unstable[3])
    

def saveresult(tprint, c_central, c_upstream, method):

    pos_anal, c_anal = [], []

    # load the analytical solution
    # T = 50 days
    wbsheet = load_workbook('cp2_t_50_3rd.xlsx')['Sheet1'].columns
    pos_anal.append([cell.value for cell in wbsheet[1][1:]])
    c_anal.append([cell.value for cell in wbsheet[2][1:]])

    # T = 100 days
    wbsheet = load_workbook('cp2_t_100_3rd.xlsx')['Sheet1'].columns
    pos_anal.append([cell.value for cell in wbsheet[1][1:]])
    c_anal.append([cell.value for cell in wbsheet[2][1:]])

    for i in xrange(len(tprint)):
        t = tprint[i]
        c_anal_v = c_anal[i]
        c_central_v = c_central[i]
        c_upstream_v = c_upstream[i]
        with open(method + "_T" + str(t) + "_values.csv", 'w') as csvfile:
            fieldnames = ['location', 'analytical', 'central', 'upstream']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            for j in xrange(len(x_pos)):
                writer.writerow({
                    'location':      '%.2f' % x_pos[j],
                    'analytical':    '%.5f' % c_anal_v[j],
                    'central':       '%.5f' % c_central_v[j],
                    'upstream':      '%.5f' % c_upstream_v[j]
                })

if __name__ == "__main__":
    main()

        # # plot if t is in tprint
        # # and check the mass balance
        # if ifin(t, tprint):
        #     # calculate the analytical solution
        #     # c_anal = analytical_c(x_pos, t)
        #     c_anal = np.zeros(N)
        #     # calculate the difference
        #     error = c_anal - c_new
        #     total_error = np.sqrt(np.mean(np.square(error)))
        #     # plot
        #     fig, ax = plt.subplots()
        #     ax.plot(x_pos, c_new, linstyles[0], color=colors[0],
        #             label="numerical")
        #     ax.plot(x_pos, c_anal, linstyles[1], color=colors[1],
        #             label="analytical")
        #     # plot setting
        #     ax.set_ylabel('c(x, t)')
        #     ax.set_xlabel('x')
        #     ax.set_ylim(0, 1)
        #     ax.set_title('%s, Script_U = %s, Script_D = %s, delta_x = %s, t = %s' %
        #                  (method_name, str(SU), str(SD), str(dx), str(t)))
        #     ax.legend(loc='lower center', shadow=True)
        #     filename = time_method + "_" + space_method + "_SD_" + str(SD) + "&dx_" + str(dx) + "&t_" + str(t)
        #     plt.savefig( filename + ".png")

        #     # check the mass balance
        #     # assuming the cross section area A = 1
        #     # also, we know that D = 1
        #     m_change, m_out = 0, 0
        #     for i in range(N):
        #         m_change += dx * ((c_old[i+1]-c_new[i+1]) +
        #                           (c_old[i]-c_new[i]))/2
        #     m_out = dt * 1 * ((c_new[1]-c_new[0])/dx +
        #                       (c_new[N-1]-c_new[N])/dx)
        #     diff = m_change - m_out
        #     print "*** Check the mass balance per unit area of the cross-section ***"
        #     print "Method: %s" % method_name
        #     print "dx: %.2f; dt: %.2f" % (dx, dt)
        #     print "Time: %s (s): " % str(t)
        #     print "Change in Storage: %.5f; Flux out of the domain: %.5f" \
        #         % (m_change, m_out)
        #     # write into the csvfile (writer)
        #     writer.writerow({'method':                 method_name,
        #                      'dt':                     "%.2f" % dt,
        #                      'dx':                     "%.2f" % dx,
        #                      't':                      "%.2f" % t,
        #                      'change in storage':      "%.5f" % m_change,
        #                      'flux out of the domain': "%.5f" % m_out,
        #                      'diff':                   "%.5f" % diff,
        #                      'total_error':            "%.5f" % total_error})

        #     # # write the h into a csv file
        #     # with open(filename + '.csv', 'w') as csvfile2:
        #     #     fieldnames2 = ['x','analytical', 'numerical', 'error']
        #     #     writer2 = csv.DictWriter(csvfile2, fieldnames=fieldnames2)
        #     #     writer2.writeheader()
        #     #     for i in xrange(N):
        #     #         writer2.writerow({'x':          "%.2f" % x_pos[i],
        #     #                           'analytical': "%.7f" % h_anal[i],
        #     #                           'numerical':  "%.7f" % h_new_f[i],
        #     #                           'error':      "%.7f" % error[i]})
